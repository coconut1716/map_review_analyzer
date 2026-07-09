from __future__ import annotations

import argparse
import csv
import json
import math
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlencode
from urllib.error import HTTPError
from urllib.request import Request, urlopen


API_BASE = "https://dapi.kakao.com"
ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT_DIR = ROOT / "outputs" / "kakao_restaurants"
EARTH_RADIUS_M = 6371000.0


def request_json(path: str, params: dict[str, str | int], api_key: str) -> dict:
    url = f"{API_BASE}{path}?{urlencode(params)}"
    req = Request(
        url,
        headers={
            "Authorization": f"KakaoAK {api_key}",
            "Accept": "application/json",
            "User-Agent": "kakao-restaurant-search/1.0",
        },
    )
    try:
        with urlopen(req, timeout=20) as res:
            return json.loads(res.read().decode("utf-8"))
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Kakao API 요청 실패: HTTP {exc.code}\nURL: {url}\n응답: {body}") from exc


def search_keyword(api_key: str, query: str, max_pages: int, size: int, delay: float) -> list[dict]:
    rows = []
    for page in range(1, max_pages + 1):
        payload = request_json(
            "/v2/local/search/keyword.json",
            {"query": query, "page": page, "size": size},
            api_key,
        )
        rows.extend(payload.get("documents", []))
        if payload.get("meta", {}).get("is_end", True):
            break
        time.sleep(delay)
    return rows


def resolve_center(api_key: str, region: str) -> tuple[str, str] | None:
    payload = request_json(
        "/v2/local/search/address.json",
        {"query": region, "size": 1},
        api_key,
    )
    docs = payload.get("documents", [])
    if docs:
        return docs[0]["x"], docs[0]["y"]

    payload = request_json(
        "/v2/local/search/keyword.json",
        {"query": region, "size": 1},
        api_key,
    )
    docs = payload.get("documents", [])
    if docs:
        return docs[0]["x"], docs[0]["y"]
    return None



def to_distance(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 999999.0


def offset_point(lon: float, lat: float, east_m: float, north_m: float) -> tuple[float, float]:
    lat_rad = math.radians(lat)
    new_lat = lat + (north_m / 111_320.0)
    new_lon = lon + (east_m / (111_320.0 * math.cos(lat_rad)))
    return new_lon, new_lat


def haversine_m(lon1: float, lat1: float, lon2: float, lat2: float) -> float:
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lon2 - lon1)
    a = math.sin(d_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    return 2 * EARTH_RADIUS_M * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def generate_grid(anchor_x: float, anchor_y: float, outer_radius: int, grid_step: int) -> list[dict]:
    points = []
    point_no = 0
    max_offset = int(math.ceil(outer_radius / grid_step) * grid_step)
    for north_m in range(-max_offset, max_offset + 1, grid_step):
        for east_m in range(-max_offset, max_offset + 1, grid_step):
            offset_distance = math.hypot(east_m, north_m)
            if offset_distance > outer_radius:
                continue
            point_no += 1
            x, y = offset_point(anchor_x, anchor_y, east_m, north_m)
            points.append(
                {
                    "grid_id": point_no,
                    "x": x,
                    "y": y,
                    "east_m": east_m,
                    "north_m": north_m,
                    "offset_distance_m": round(offset_distance, 1),
                }
            )
    return points


def distance_bucket(distance_m: float, cuts: list[int]) -> str:
    low = 0
    for cut in cuts:
        if distance_m <= cut:
            return f"{low}-{cut}m"
        low = cut
    return f">{cuts[-1]}m"
def search_food_category(
    api_key: str,
    x: str,
    y: str,
    radius: int,
    max_pages: int,
    size: int,
    delay: float,
) -> list[dict]:
    rows = []
    for page in range(1, max_pages + 1):
        payload = request_json(
            "/v2/local/search/category.json",
            {
                "category_group_code": "FD6",
                "x": x,
                "y": y,
                "radius": radius,
                "page": page,
                "size": size,
                "sort": "distance",
            },
            api_key,
        )
        rows.extend(payload.get("documents", []))
        if payload.get("meta", {}).get("is_end", True):
            break
        time.sleep(delay)
    return rows


def normalize(doc: dict, region: str, source: str) -> dict:
    place_url = doc.get("place_url", "")
    review_url = f"{place_url.split('#', 1)[0]}#review" if place_url else ""
    return {
        "id": doc.get("id", ""),
        "place_name": doc.get("place_name", ""),
        "category_name": doc.get("category_name", ""),
        "category_group_name": doc.get("category_group_name", ""),
        "phone": doc.get("phone", ""),
        "address_name": doc.get("address_name", ""),
        "road_address_name": doc.get("road_address_name", ""),
        "x": doc.get("x", ""),
        "y": doc.get("y", ""),
        "distance": doc.get("distance", ""),
        "place_url": place_url,
        "review_url": review_url,
        "region_query": region,
        "source": source,
    }



def normalize_grid(doc: dict, region: str, anchor: dict, source_grid: dict, bucket_cuts: list[int]) -> dict:
    row = normalize(doc, region, "grid")
    x = float(doc["x"]) if doc.get("x") else 0.0
    y = float(doc["y"]) if doc.get("y") else 0.0
    anchor_distance = haversine_m(anchor["x"], anchor["y"], x, y)
    row.update(
        {
            "distance": f"{anchor_distance:.1f}",
            "anchor_distance_m": round(anchor_distance, 1),
            "distance_bucket": distance_bucket(anchor_distance, bucket_cuts),
            "grid_id": source_grid["grid_id"],
            "grid_offset_east_m": source_grid["east_m"],
            "grid_offset_north_m": source_grid["north_m"],
            "grid_distance_m": doc.get("distance", ""),
        }
    )
    return row


def collect_grid_category(
    api_key: str,
    region: str,
    outer_radius: int,
    grid_step: int,
    cell_radius: int,
    max_pages: int,
    size: int,
    delay: float,
    bucket_cuts: list[int],
) -> list[dict]:
    center = resolve_center(api_key, region)
    if not center:
        print(f"기준 좌표를 찾지 못해 grid 검색을 건너뜁니다: {region}", file=sys.stderr)
        return []

    anchor_x, anchor_y = map(float, center)
    anchor = {"x": anchor_x, "y": anchor_y}
    grid_points = generate_grid(anchor_x, anchor_y, outer_radius, grid_step)
    print(f"grid 검색 기준 좌표: x={anchor_x:.8f}, y={anchor_y:.8f}")
    print(f"grid 검색: outer={outer_radius}m, step={grid_step}m, cell={cell_radius}m, points={len(grid_points)}")

    collected = []
    for idx, point in enumerate(grid_points, start=1):
        docs = search_food_category(
            api_key,
            f"{point['x']:.8f}",
            f"{point['y']:.8f}",
            cell_radius,
            max_pages,
            size,
            delay,
        )
        rows = [normalize_grid(doc, region, anchor, point, bucket_cuts) for doc in docs]
        kept = [row for row in rows if to_distance(row.get("anchor_distance_m")) <= outer_radius]
        collected.extend(kept)
        print(f"  grid {idx:03d}/{len(grid_points)} raw={len(docs)} kept={len(kept)}")
        time.sleep(delay)
    return collected
def dedupe(rows: list[dict]) -> list[dict]:
    out = {}
    for row in rows:
        key = row.get("id") or f"{row.get('place_name')}|{row.get('address_name')}"
        old = out.get(key)
        if old is None or to_distance(row.get("distance")) < to_distance(old.get("distance")):
            out[key] = row
    return sorted(out.values(), key=lambda r: (to_distance(r.get("distance")), r.get("place_name", "")))


def save_csv(path: Path, rows: list[dict], fields: list[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def save_xlsx(path: Path, rows: list[dict], fields: list[str]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    ws = wb.active
    ws.title = "restaurants"
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"

    widths = {
        "place_name": 26,
        "category_name": 34,
        "phone": 16,
        "address_name": 34,
        "road_address_name": 38,
        "place_url": 38,
        "review_url": 44,
    }
    for idx, field in enumerate(fields, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(field, 16)

    wb.save(path)
    return True


def main() -> int:
    parser = argparse.ArgumentParser(description="Kakao Local API로 특정 지역 음식점 목록과 리뷰 링크를 저장합니다.")
    parser.add_argument("region", help="검색할 지역. 예: '인천 연수구 송도동'")
    parser.add_argument("--api-key", default=os.getenv("KAKAO_REST_API_KEY"), help="Kakao REST API 키. 기본값: KAKAO_REST_API_KEY")
    parser.add_argument("--mode", choices=["keyword", "category", "both", "grid"], default="both")
    parser.add_argument("--radius", type=int, default=2000, help="category 모드 반경(m), 최대 20000")
    parser.add_argument("--outer-radius", type=int, default=300, help="grid 모드 최종 기준점 반경(m)")
    parser.add_argument("--grid-step", type=int, default=100, help="grid 모드 검색 중심 좌표 간격(m)")
    parser.add_argument("--cell-radius", type=int, default=120, help="grid 모드 각 격자점 검색 반경(m)")
    parser.add_argument("--bucket-cuts", default="250,350,430,500", help="거리 구간 표시용 쉼표 구분 컷(m)")
    parser.add_argument("--minimal", action="store_true", help="식당 목록 CSV/XLSX에 핵심 필드만 저장")
    parser.add_argument("--max-pages", type=int, default=45)
    parser.add_argument("--size", type=int, default=15, help="페이지당 개수, 최대 15")
    parser.add_argument("--delay", type=float, default=0.2, help="페이지 요청 사이 대기초")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    if not args.api_key:
        print("KAKAO_REST_API_KEY 환경변수 또는 --api-key가 필요합니다.", file=sys.stderr)
        return 2
    if not (1 <= args.size <= 15):
        print("--size는 1~15 사이여야 합니다.", file=sys.stderr)
        return 2
    if not (1 <= args.radius <= 20000):
        print("--radius는 1~20000 사이여야 합니다.", file=sys.stderr)
        return 2
    if not (1 <= args.outer_radius <= 20000):
        print("--outer-radius는 1~20000 사이여야 합니다.", file=sys.stderr)
        return 2
    if args.grid_step <= 0 or args.cell_radius <= 0:
        print("--grid-step과 --cell-radius는 양수여야 합니다.", file=sys.stderr)
        return 2
    bucket_cuts = [int(part.strip()) for part in args.bucket_cuts.split(",") if part.strip()] or [args.outer_radius]

    collected = []
    if args.mode in ("keyword", "both"):
        query = f"{args.region} 음식점"
        collected.extend(normalize(doc, args.region, "keyword") for doc in search_keyword(args.api_key, query, args.max_pages, args.size, args.delay))

    if args.mode == "grid":
        collected.extend(
            collect_grid_category(
                args.api_key,
                args.region,
                args.outer_radius,
                args.grid_step,
                args.cell_radius,
                args.max_pages,
                args.size,
                args.delay,
                bucket_cuts,
            )
        )

    if args.mode in ("category", "both"):
        center = resolve_center(args.api_key, args.region)
        if center:
            x, y = center
            collected.extend(
                normalize(doc, args.region, "category")
                for doc in search_food_category(args.api_key, x, y, args.radius, args.max_pages, args.size, args.delay)
            )
        else:
            print(f"지역 중심 좌표를 찾지 못해 category 검색을 건너뜁니다: {args.region}", file=sys.stderr)

    rows = dedupe(collected)
    now = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    safe_region = "".join(ch if ch.isalnum() else "_" for ch in args.region).strip("_")
    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    base = out_dir / f"kakao_restaurants_{safe_region}_{now}"

    fields = [
        "id",
        "place_name",
        "category_name",
        "category_group_name",
        "phone",
        "address_name",
        "road_address_name",
        "x",
        "y",
        "distance",
        "place_url",
        "review_url",
        "region_query",
        "source",
        "anchor_distance_m",
        "distance_bucket",
        "grid_id",
        "grid_offset_east_m",
        "grid_offset_north_m",
        "grid_distance_m",
    ]
    fields_minimal = ["id", "place_name", "category_name", "distance", "anchor_distance_m", "distance_bucket", "review_url"]
    if args.minimal:
        fields = fields_minimal

    json_path = base.with_suffix(".json")
    csv_path = base.with_suffix(".csv")
    xlsx_path = base.with_suffix(".xlsx")

    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    save_csv(csv_path, rows, fields)
    xlsx_ok = save_xlsx(xlsx_path, rows, fields)

    print(f"수집 식당 수: {len(rows)}")
    print(f"JSON: {json_path}")
    print(f"CSV:  {csv_path}")
    if xlsx_ok:
        print(f"XLSX: {xlsx_path}")
    else:
        print("XLSX: openpyxl이 없어 생략했습니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

